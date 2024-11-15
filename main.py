import streamlit as st
import pdfplumber
from bidi.algorithm import get_display
import arabic_reshaper
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
import io

GUSH_HELKA_BANNER_LINE = 7

def parse_banner(line):
    line = line.split()
    return {
        "gush": line[1],
        "plot": line[3]
    }

def extract_data_from_subplot(lines):
    subplot = lines[0]
    subplot_info = lines[2].split()
    area_in_square_meters = subplot_info[0]
    floor_description = subplot_info[1]
    share_in_the_common_property = subplot_info[2]

    ownerships = ""
    linkage = ""
    mortage = ""
    lease = ""
    notes = ""

    start_info_section_id = 3
    for i in range(4, len(lines)):
        if lines[i] in ['בעלויות', 'משכנתאות', 'הצמדות', 'חכירות', 'הערות'] or i == len(lines) - 1:
            info_section = lines[start_info_section_id:i]
            start_info_section_id = i
            if info_section:
                section_title = info_section[0]
                content = info_section[1:]
                if section_title == 'בעלויות':
                    ownerships = content
                elif section_title == 'משכנתאות':
                    mortage = content
                elif section_title == 'הצמדות':
                    linkage = [line for line in content if line != 'סימון בתשריט צבע בתשריט תיאור הצמדה שטח במ"ר']
                elif section_title == 'חכירות':
                    lease = content
                elif section_title == 'הערות':
                    notes = content
    return {
        "subplot": subplot,
        "area_in_square_meters": area_in_square_meters,
        "floor_description": floor_description,
        "share_in_the_common_property": share_in_the_common_property,
        "ownerships": ownerships,
        "linkage": linkage,
        "mortage": mortage,
        "lease": lease,
        "notes": notes,
    }

def split_subplot(lines):
    # Find first subplot
    first_subplot_id = next((i for i in range(len(lines)) if lines[i].startswith('תת חלקה')), None)
    if first_subplot_id is None:
        return []

    # Find end of data
    end_subplot_id = next((i for i in range(len(lines)) if lines[i] == 'סוף נתונים'), len(lines))

    lines = lines[first_subplot_id:end_subplot_id]

    # Split by subplots
    subplots = []
    start_id = 0
    for i in range(1, len(lines)):
        if lines[i].startswith('תת חלקה '):
            subplots.append(lines[start_id:i])
            start_id = i

    subplots.append(lines[start_id:])
    return subplots

def extract_tables_from_pdf(pdf_file):
    all_text = []
    plot_info = {"gush": "", "plot": ""}

    # Open the PDF
    with pdfplumber.open(pdf_file) as pdf:

        # Iterate through each page
        for page_number, page in enumerate(pdf.pages, start=1):

            # Process each page
            data = page.extract_text()
            if not data:
                continue  # Skip pages with no text

            data = data.splitlines()
            data = [get_display(arabic_reshaper.reshape(line)) for line in data]

            if len(data) > GUSH_HELKA_BANNER_LINE:
                plot_info = parse_banner(data[GUSH_HELKA_BANNER_LINE])
                data = data[GUSH_HELKA_BANNER_LINE+1:-1]
                all_text += data
            else:
                continue  # Skip pages that don't have enough lines

    return all_text, plot_info

def create_excel(plot_info, lines):
    for i in range(len(lines)):
        for key in lines[i].keys():
            if isinstance(lines[i][key], list):
                lines[i][key] = '\n'.join(lines[i][key])

    df = pd.DataFrame(lines)
    df['gush'] = plot_info['gush']
    df['plot'] = plot_info['plot']

    # Rearrange columns order
    df = df[['gush', 'plot','subplot', 'area_in_square_meters', 'floor_description',
       'share_in_the_common_property', 'ownerships', 'linkage',
       'mortage', 'lease', 'notes']]

    # Rename columns
    df = df.rename(columns={
        'gush': 'גוש',
        'plot': 'חלקה',
        'subplot': 'תת חלקה',
        'area_in_square_meters': 'שטח במ"ר',
        'floor_description': 'תיאור קומה',
        'share_in_the_common_property': 'החלק ברכוש המשותף',
        'ownerships': 'בעלויות',
        'linkage': 'הצמדות',
        'mortage': 'משכנתאות',
        'lease': 'חכירות',
        'notes': 'הערות'
    })

    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active

    # Set the sheet to Right-to-Left (RTL)
    ws.sheet_view.rightToLeft = True

    # Write the data from the DataFrame to the worksheet
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    # Apply styles to the header
    header_fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Apply alignment to all cells and wrap text
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    # Automatically adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column  # Get the column number
        column_letter = get_column_letter(column)  # Convert to letter
        for cell in col:
            try:
                if cell.value:  # Ensure the cell is not None
                    max_length = max(max_length, max([len(x) for x in str(cell.value).split('\n')]))
            except:
                pass
        adjusted_width = max_length + 2  # Add some padding
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save the workbook to a BytesIO object
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

def main():
    # Set page configuration
    st.set_page_config(
        page_title="PDF to Excel Converter",
        page_icon=":page_facing_up:",
        layout="centered"
    )

    # Custom CSS for styling
    st.markdown("""
        <style>
            /* Center everything */
            .block-container {
                max-width: 800px;
                padding-top: 2rem;
                padding-bottom: 2rem;
            }
            /* Style the upload button */
            .stFileUploader > label {
                text-align: center;
                display: block;
            }
            /* Style the download button */
            .stDownloadButton > button {
                color: white;
                background-color: #4CAF50;
            }
            /* Footer */
            .footer {
                position: fixed;
                left: 0;
                bottom: 0;
                width: 100%;
                text-align: center;
                color: gray;
            }
            /* Hide Streamlit header and footer */
            #MainMenu {visibility: hidden;}
            footer {visibility: hidden;}
        </style>
        """, unsafe_allow_html=True)


    st.title("Tabu PDF to Excel Converter")
    st.write("Easily convert your PDF files into Excel spreadsheets.")

    # Initialize session state
    if 'excel_data' not in st.session_state:
        st.session_state['excel_data'] = None
    if 'processing_done' not in st.session_state:
        st.session_state['processing_done'] = False
    if 'uploaded_filename' not in st.session_state:
        st.session_state['uploaded_filename'] = None

    # File uploader
    uploaded_file = st.file_uploader("Please upload a PDF file:", type=["pdf"])

    if uploaded_file is not None:
        # Check if we have already processed this file
        if st.session_state['uploaded_filename'] != uploaded_file.name:
            # Reset session state
            st.session_state['excel_data'] = None
            st.session_state['processing_done'] = False
            st.session_state['uploaded_filename'] = uploaded_file.name

        if not st.session_state['processing_done']:
            # Show progress bar
            progress_bar = st.progress(0)
            status_text = st.empty()

            try:
                # Extract all the text
                progress_bar.progress(10)
                status_text.text("Extracting text from PDF...")
                all_text, plot_info = extract_tables_from_pdf(uploaded_file)

                if not all_text:
                    st.warning("No text found in the PDF file.")
                    progress_bar.empty()
                    status_text.empty()
                    return

                progress_bar.progress(50)
                status_text.text("Parsing subplots...")

                subplots_parsed = []
                # Parse each subplot
                for subplot in split_subplot(all_text):
                    line = extract_data_from_subplot(subplot)
                    subplots_parsed.append(line)

                if not subplots_parsed:
                    st.warning("No data found in the PDF file.")
                    progress_bar.empty()
                    status_text.empty()
                    return

                # Save to Excel
                progress_bar.progress(80)
                status_text.text("Creating Excel file...")

                excel_data = create_excel(plot_info, subplots_parsed)

                progress_bar.progress(100)
                status_text.text("Processing complete!")

                # Store the excel_data in session state
                st.session_state['excel_data'] = excel_data
                st.session_state['processing_done'] = True

                progress_bar.empty()
                status_text.empty()

            except Exception as e:
                st.error("An error occurred during processing.")
                st.exception(e)
                progress_bar.empty()
                status_text.empty()
                return

        if st.session_state['processing_done']:
            # Download button with custom text
            st.success("✅ Your Excel file is ready for download.")
            st.download_button(
                label="📥 Download Excel file",
                data=st.session_state['excel_data'],
                file_name="output.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.info("Please upload a PDF file to begin.")

    # Footer with creator's name and GitHub link
    st.markdown("""
        <div class="footer">
            <hr>
            Created by <a href="https://github.com/Boris-Rozenfeld" target="_blank">Boris Rozenfeld</a> |
            <a href="https://github.com/Boris-Rozenfeld/Tabu-PDF-To-Excel-Conventor/issues" target="_blank">Report an Issue</a>
        </div>
        """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()
